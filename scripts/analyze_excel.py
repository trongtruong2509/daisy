#!/usr/bin/env python3
"""
Script to analyze the Excel file structure for payslip tool requirements.
"""

import sys
from pathlib import Path

# Try different Excel libraries
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not available")

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    print("xlrd not available")

def analyze_with_openpyxl(file_path):
    """Analyze Excel file using openpyxl."""
    print(f"\n{'='*60}")
    print(f"Analyzing with openpyxl: {file_path}")
    print('='*60)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f"\n📋 Sheet Names ({len(wb.sheetnames)} sheets):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # Analyze each important sheet
        for sheet_name in wb.sheetnames:
            print(f"\n{'─'*60}")
            print(f"Sheet: {sheet_name}")
            print('─'*60)
            
            ws = wb[sheet_name]
            print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Show first few rows
            if ws.max_row > 0:
                print(f"\nFirst 5 rows:")
                for row_idx in range(1, min(6, ws.max_row + 1)):
                    row_data = []
                    for col_idx in range(1, min(11, ws.max_column + 1)):  # First 10 columns
                        cell = ws.cell(row_idx, col_idx)
                        value = cell.value
                        if value is not None:
                            row_data.append(str(value)[:30])  # Truncate long values
                        else:
                            row_data.append("")
                    print(f"  Row {row_idx}: {' | '.join(row_data)}")
                
                # If this is the Data sheet, find all column headers
                if sheet_name.lower() in ['data', 'bang luong']:
                    print(f"\n📊 Column Headers (Row 3):")
                    headers = []
                    for col_idx in range(1, ws.max_column + 1):
                        header_cell = ws.cell(3, col_idx)
                        if header_cell.value:
                            # Get column letter
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            headers.append(f"{col_letter}: {header_cell.value}")
                    
                    for i, header in enumerate(headers, 1):
                        print(f"  {i}. {header}")
                        if i > 30:  # Limit output
                            print(f"  ... and {len(headers) - 30} more columns")
                            break
    
    except Exception as e:
        print(f"Error analyzing with openpyxl: {e}")
        import traceback
        traceback.print_exc()

def analyze_with_xlrd(file_path):
    """Analyze Excel file using xlrd (for .xls files)."""
    print(f"\n{'='*60}")
    print(f"Analyzing with xlrd: {file_path}")
    print('='*60)
    
    try:
        wb = xlrd.open_workbook(file_path)
        print(f"\n📋 Sheet Names ({wb.nsheets} sheets):")
        for i, sheet in enumerate(wb.sheets()):
            print(f"  {i+1}. {sheet.name}")
        
        # Analyze each sheet
        for sheet in wb.sheets():
            print(f"\n{'─'*60}")
            print(f"Sheet: {sheet.name}")
            print('─'*60)
            print(f"Dimensions: {sheet.nrows} rows x {sheet.ncols} columns")
            
            if sheet.nrows > 0:
                print(f"\nFirst 5 rows:")
                for row_idx in range(min(5, sheet.nrows)):
                    row_data = []
                    for col_idx in range(min(10, sheet.ncols)):
                        cell = sheet.cell(row_idx, col_idx)
                        value = str(cell.value)[:30] if cell.value else ""
                        row_data.append(value)
                    print(f"  Row {row_idx+1}: {' | '.join(row_data)}")
                
                # If this is the Data sheet, find all column headers
                if sheet.name.lower() in ['data', 'bang luong']:
                    print(f"\n📊 Column Headers (Row 3):")
                    if sheet.nrows >= 3:
                        headers = []
                        for col_idx in range(sheet.ncols):
                            header_value = sheet.cell(2, col_idx).value  # Row 3 is index 2
                            if header_value:
                                # Convert column index to letter
                                col_letter = xlrd.colname(col_idx)
                                headers.append(f"{col_letter}: {header_value}")
                        
                        for i, header in enumerate(headers, 1):
                            print(f"  {i}. {header}")
                            if i > 30:
                                print(f"  ... and {len(headers) - 30} more columns")
                                break
    
    except Exception as e:
        print(f"Error analyzing with xlrd: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    file_path = Path(__file__).parent / "excel-files" / "TBKQ-phuclong.xls"
    
    if not file_path.exists():
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    print(f"📁 File: {file_path}")
    print(f"📦 File size: {file_path.stat().st_size / 1024:.2f} KB")
    
    # Try xlrd first for .xls files
    if HAS_XLRD:
        analyze_with_xlrd(file_path)
    elif HAS_OPENPYXL:
        analyze_with_openpyxl(file_path)
    else:
        print("❌ No Excel library available. Install with: pip install xlrd openpyxl")
        sys.exit(1)
