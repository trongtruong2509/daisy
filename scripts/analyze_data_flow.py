#!/usr/bin/env python3
"""
Analyze what data needs to be populated in TBKQ template.
Understand the Data sheet structure and how it maps to TBKQ cells.
"""

from openpyxl import load_workbook

excel_file = "excel-files/TBKQ-phuclong.xls"

print("=" * 100)
print("UNDERSTANDING DATA FLOW: Data Sheet → TBKQ Template")
print("=" * 100)

# Open with openpyxl to see if any formulas are preserved
try:
    wb = load_workbook(excel_file, data_only=False)
    
    # Get Data sheet
    data_sheet = wb['Data']
    headers = []
    
    print("\n1. DATA SHEET STRUCTURE (Headers in Row 2)")
    print("-" * 100)
    print("This is the source data that needs to be populated into TBKQ template\n")
    
    # Read headers from row 2
    for col_idx in range(1, data_sheet.max_column + 1):
        header = data_sheet.cell(2, col_idx).value
        if header:
            # Convert column number to letter
            col_letter = ''
            n = col_idx
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                col_letter = chr(65 + remainder) + col_letter
            
            headers.append({
                'letter': col_letter,
                'index': col_idx,
                'name': str(header).strip()
            })
    
    for h in headers[:20]:  # Show first 20 columns
        print(f"  Column {h['letter']:3s} ({h['index']:2d}): {h['name']}")
    
    print(f"\n  ... and {len(headers) - 20} more columns (total {len(headers)} columns)")
    
    print("\n2. SAMPLE DATA (Row 4 - First Employee)")
    print("-" * 100)
    for h in headers[:10]:
        value = data_sheet.cell(4, h['index']).value
        print(f"  {h['letter']}: {h['name']:30s} = {value}")
    
    # Now check TBKQ for formulas
    print("\n3. TBKQ TEMPLATE ANALYSIS - Looking for Formulas")
    print("-" * 100)
    
    tbkq_sheet = wb['TBKQ']
    
    formula_cells = []
    for row in tbkq_sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=11):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_cells.append({
                    'cell': f"{cell.column_letter}{cell.row}",
                    'formula': cell.value
                })
    
    if formula_cells:
        print(f"\nFound {len(formula_cells)} formulas in TBKQ sheet:\n")
        for fc in formula_cells[:10]:
            print(f"  {fc['cell']}: {fc['formula']}")
        
        print("\nThese formulas need to be replaced with CALCULATED VALUES from Data sheet")
    else:
        print("\nNo formulas found in .xls file (formulas may have been lost during conversion)")
        print("This means we need to:")
        print("  1. Analyze the template structure manually")
        print("  2. Understand what each cell should contain")
        print("  3. Create mapping: TBKQ cell → Data sheet column")
    
    wb.close()
    
except Exception as e:
    print(f"Error: {e}")

print("\n" + "=" * 100)
print("SOLUTION FOR DIRECT POPULATION")
print("=" * 100)

print("""
STEP 1: Create a Cell Mapping Configuration
=========================================

For each TBKQ cell, define which Data column to pull from:

mapping = {
    'B3': 'A',      # Employee MNV from Data column A
    'B4': 'B',      # Employee Name from Data column B
    'D3': 'C',      # Start Date from Data column C
    'B9': 'D',      # Salary from Data column D
    'D9': 'E',      # Total salary from Data column E
    # ... map all cells that need data ...
}

STEP 2: For Each Employee, Execute This Logic
=============================================

def populate_payslip(employee_row_number, data_sheet, template_sheet, mapping):
    
    # Read all employee data from Data sheet
    employee_data = {}
    for col_letter, data_col in mapping.items():
        value = data_sheet.cell(employee_row_number, col=column_to_number(data_col)).value
        employee_data[col_letter] = value
    
    # Copy template to new workbook
    new_workbook = copy_template()
    tbkq = new_workbook['TBKQ']
    
    # Write values (not formulas) to TBKQ cells
    for tbkq_cell, data_col in mapping.items():
        tbkq[tbkq_cell].value = employee_data[tbkq_cell]  # Direct value, not formula!
    
    # Save to PDF
    new_workbook.save('temp.xlsx')
    convert_xlsx_to_pdf('temp.xlsx', 'payslip.pdf')

STEP 3: What About Calculated Fields?
====================================

Some TBKQ cells might contain FORMULAS that CALCULATE values (like SUM):

    D9 = SUM(D8:D8)  →  Should become: D9 = 1500  (the calculated result)

Solution: When copying template, we need to:
  1. Get the CALCULATED VALUE (not the formula) from the original template
  2. Or calculate it ourselves in Python
  3. Write the result value (not formula) to the new payslip

EXAMPLE:
--------
If Data sheet has raw salary components in columns D, E, F, G...
And TBKQ template has a formula to sum them: =D10+E10+F10

We can:
  Option A: Read the formula result from original file
  Option B: Calculate sum in Python: value = D + E + F, write to TBKQ
""")

print("\n" + "=" * 100)
print("CRITICAL NEXT STEP: GET THE ORIGINAL .XLSX FILE")
print("=" * 100)

print("""
To properly implement direct population, we NEED:

1. The original TBKQ-phuclong.xlsx file (with formulas intact)
   - This shows us which cells contain formulas
   - Which formulas reference which sheets/columns
   - What the expected output structure is

2. OR: VBA source code or documentation
   - Shows exactly which cells are populated and from where
   - Defines the mapping between Data sheet and TBKQ template

WITHOUT this, we're guessing about cell mappings!

Question for user:
Do you have the original .xlsx file or VBA source with cell mapping info?
""")
