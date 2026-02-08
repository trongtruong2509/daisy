"""
Generate batch test data for payslip tool.

Copies TBKQ-phuclong.xls to a new test file and populates it with 200+ employees.
Employee emails are randomly assigned between two addresses for testing multiple account sends.
"""

import random
import sys
import shutil
from pathlib import Path

# Try openpyxl first (works with .xlsx), then fallback to other methods
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import win32com.client as win32
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False

if not HAS_OPENPYXL and not HAS_WIN32COM:
    print("ERROR: Neither openpyxl nor pywin32 available.")
    print("Install one with: pip install openpyxl  or  pip install pywin32")
    sys.exit(1)


# Vietnamese names for generating realistic test data
FIRST_NAMES = [
    "Nguyễn", "Trần", "Hoàng", "Phạm", "Huỳnh", "Đinh", "Bùi", "Đặng", "Vũ", "Võ",
    "Dương", "Lý", "Phan", "Tô", "Tạ", "Lê", "Đỗ", "Đào", "Từ", "Lâm"
]

MIDDLE_NAMES = [
    "Văn", "Thái", "Hữu", "Huy", "Hoài", "Kiến", "Minh", "Thanh", "Hùng", "Tiến",
    "Đình", "Anh", "Hải", "Duy", "Phong", "Tuấn", "Trí", "Quang", "Nhân", "Đức"
]

LAST_NAMES = [
    "A", "B", "C", "D", "E", "Sơn", "Hải", "Hào", "Hân", "Hiền",
    "Hồng", "Hương", "Huy", "Huỳnh", "Huỳnh", "Hùng", "Hạnh", "Hải", "Hân", "Hiền"
]

TEST_EMAILS = ["tht.tts@gmail.com", "trong.h.truong@gmail.com"]


def generate_vietnamese_name():
    """Generate a random Vietnamese name."""
    first = random.choice(FIRST_NAMES)
    middle = random.choice(MIDDLE_NAMES)
    last = random.choice(LAST_NAMES)
    return f"{first} {middle} {last}"


def generate_with_openpyxl(source_file: Path, output_file: Path, num_employees: int = 200):
    """Generate batch file using openpyxl (for .xlsx files)."""
    print(f"[1/3] Copying Excel file...")
    shutil.copy2(source_file, output_file)
    print(f"      ✓ Copied")
    
    print(f"[2/3] Loading and populating with openpyxl...")
    wb = openpyxl.load_workbook(output_file)
    
    try:
        ws = wb["Data"]
    except KeyError:
        print(f"      ✗ Could not find 'Data' sheet")
        return False
    
    # Column positions
    col_a, col_b, col_c = 1, 2, 3  # A, B, C
    col_az = 52  # AZ
    start_row = 4
    
    # Get existing password
    existing_password = ws.cell(start_row, col_az).value
    print(f"      - Existing password sample: {existing_password}")
    
    # Generate data
    mnv_counter = 6046073
    for i in range(num_employees):
        row = start_row + i
        ws.cell(row, col_a).value = mnv_counter
        ws.cell(row, col_b).value = generate_vietnamese_name()
        ws.cell(row, col_c).value = random.choice(TEST_EMAILS)
        ws.cell(row, col_az).value = existing_password
        mnv_counter += 1
        
        if (i + 1) % 50 == 0:
            print(f"      - Generated {i + 1}/{num_employees}...")
    
    print(f"[3/3] Saving file...")
    wb.save(output_file)
    wb.close()
    
    # Calculate stats
    emails = [ws.cell(start_row + i, col_c).value for i in range(num_employees)]
    email_count = {email: emails.count(email) for email in TEST_EMAILS}
    
    print(f"      ✓ Saved")
    print(f"      - Email distribution:")
    for email, count in email_count.items():
        pct = (count / num_employees) * 100
        print(f"        • {email}: {count} ({pct:.1f}%)")
    
    return True


def generate_with_win32com(source_file: Path, output_file: Path, num_employees: int = 200):
    """Generate batch file using Excel COM."""
    print(f"[1/4] Copying Excel file...")
    shutil.copy2(source_file, output_file)
    print(f"      ✓ Copied")
    
    print(f"[2/4] Opening Excel via COM...")
    excel = None
    workbook = None
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        workbook = excel.Workbooks.Open(str(output_file), UpdateLinks=0)
        data_sheet = workbook.Sheets("Data")
        print(f"      ✓ Opened")
    except Exception as e:
        print(f"      ✗ Failed: {e}")
        return False
    
    try:
        print(f"[3/4] Generating {num_employees} employees...")
        
        col_a, col_b, col_c, col_az = 1, 2, 3, 52
        start_row = 4
        
        existing_password = data_sheet.Cells(start_row, col_az).Value
        print(f"      - Password sample: {existing_password}")
        
        mnv_counter = 6046073
        for i in range(num_employees):
            row = start_row + i
            data_sheet.Cells(row, col_a).Value = mnv_counter
            data_sheet.Cells(row, col_b).Value = generate_vietnamese_name()
            data_sheet.Cells(row, col_c).Value = random.choice(TEST_EMAILS)
            data_sheet.Cells(row, col_az).Value = existing_password
            mnv_counter += 1
            
            if (i + 1) % 50 == 0:
                print(f"      - Generated {i + 1}/{num_employees}...")
        
        print(f"[4/4] Saving...")
        workbook.Save()
        print(f"      ✓ Saved")
        
    finally:
        try:
            workbook.Close(SaveChanges=False)
        except:
            pass
        try:
            excel.Quit()
        except:
            pass
    
    return True


def generate_batch_test_file(source_file: Path, output_file: Path, num_employees: int = 200):
    """Generate batch test file with employees."""
    if not source_file.exists():
        print(f"ERROR: Source file not found: {source_file}")
        return False
    
    # Choose method based on availability
    if HAS_OPENPYXL and str(source_file).endswith('.xlsx'):
        success = generate_with_openpyxl(source_file, output_file, num_employees)
    elif HAS_WIN32COM:
        success = generate_with_win32com(source_file, output_file, num_employees)
    else:
        print("ERROR: No suitable library available for .xls files")
        return False
    
    if success:
        print(f"\n✅ SUCCESS: {output_file}")
        print(f"   Employees: {num_employees}")
        print(f"   MNV range: 6046073-{6046073 + num_employees - 1}")
        print(f"   Emails: {', '.join(TEST_EMAILS)}")
    
    return success


if __name__ == "__main__":
    excel_files_dir = Path(__file__).parent.parent.parent.parent / "excel-files"
    source_file = excel_files_dir / "TBKQ-phuclong.xls"
    
    # Check if .xlsx version exists (better for openpyxl)
    source_xlsx = source_file.with_suffix('.xlsx')
    if source_xlsx.exists():
        source_file = source_xlsx
        output_file = excel_files_dir / "TBKQ-phuclong-batch-200.xlsx"
    else:
        output_file = excel_files_dir / "TBKQ-phuclong-batch-200.xls"
    
    print("=" * 60)
    print("BATCH TEST DATA GENERATOR")
    print("=" * 60)
    print(f"Source: {source_file}")
    print(f"Output: {output_file}")
    print(f"Using: {'openpyxl' if HAS_OPENPYXL else 'Excel COM'}")
    print("=" * 60)
    
    success = generate_batch_test_file(source_file, output_file, num_employees=210)
    sys.exit(0 if success else 1)

