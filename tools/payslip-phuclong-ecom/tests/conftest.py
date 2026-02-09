"""
Shared test fixtures for payslip tool tests.
"""

import os
import sys
from pathlib import Path

import pytest

# Set up paths
TOOL_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = TOOL_DIR.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(TOOL_DIR))

# Sample Excel file
SAMPLE_EXCEL = PROJECT_ROOT / "excel-files" / "TBKQ-phuclong.xls"


@pytest.fixture
def sample_employee():
    """A single valid employee dict."""
    return {
        "row": 4,
        "mnv": "6046072",
        "name": "Nguyen Van A",
        "email": "test@example.com",
        "password": "6046072",
        "columns": {
            "A": 6046072.0,
            "B": "Nguyen Van A",
            "C": "test@example.com",
            "F": "01/01/2020",
            "J": 15000000.0,
            "K": 26.0,
            "N": 22.0,
            "O": 12000000.0,
            "P": 2000000.0,
            "Q": 500000.0,
            "R": 300000.0,
            "S": 200000.0,
            "T": 1000000.0,
            "U": 100000.0,
            "V": 500000.0,
            "W": 0.0,
            "X": 0.0,
            "Y": 0.0,
            "Z": 200000.0,
            "AA": 150000.0,
            "AB": 1200000.0,
            "AC": 100000.0,
            "AD": 500000.0,
            "AH": 13050000.0,
            "AZ": 6046072.0,
        },
    }


@pytest.fixture
def sample_employees(sample_employee):
    """Multiple test employees."""
    emp2 = {
        "row": 5,
        "mnv": "7012345",
        "name": "Tran Thi B",
        "email": "tranb@example.com",
        "password": "7012345",
        "columns": {
            "A": 7012345.0,
            "B": "Tran Thi B",
            "C": "tranb@example.com",
            "F": "03/15/2021",
            "J": 18000000.0,
            "K": 26.0,
            "N": 24.0,
            "O": 15000000.0,
            "P": 3000000.0,
            "Q": 600000.0,
            "R": 0.0,
            "S": 0.0,
            "T": 0.0,
            "U": 0.0,
            "V": 1000000.0,
            "W": 0.0,
            "X": 0.0,
            "Y": 0.0,
            "Z": 0.0,
            "AA": 150000.0,
            "AB": 1500000.0,
            "AC": 120000.0,
            "AD": 800000.0,
            "AH": 16030000.0,
            "AZ": 7012345.0,
        },
    }
    return [sample_employee, emp2]


@pytest.fixture
def email_template():
    """Sample email template cells from bodymail."""
    return {
        "A1": "Kính gửi Anh/Chị,",
        "A3": "Công ty gửi đến Anh/Chị Thông báo phiếu lương của kỳ lương tháng 11/2025.",
        "A5": "Mật khẩu mở file: Mã số nhân viên của Anh/Chị (7 chữ số không bao gồm số 0 đầu tiên)",
        "A7": "Mọi thắc mắc liên quan, Anh/Chị vui lòng phản hồi qua email \"vn.phuclong@adecco.com\".",
        "A9": "Trân trọng,",
        "A11": "ADECCO VIỆT NAM",
        "A12": "Email: vn.phuclong@adecco.com",
    }


@pytest.fixture
def default_cell_mapping():
    """Default TBKQ cell-to-Data column mapping."""
    from config import DEFAULT_CELL_MAPPING
    return dict(DEFAULT_CELL_MAPPING)


@pytest.fixture
def default_calc_mapping():
    """Default calculated cells mapping."""
    from config import DEFAULT_CALC_MAPPING
    return dict(DEFAULT_CALC_MAPPING)


@pytest.fixture
def tmp_output(tmp_path):
    """Temporary output directory."""
    out = tmp_path / "output"
    out.mkdir()
    return out


@pytest.fixture
def minimal_template(tmp_output):
    """Create a minimal .xlsx template for testing."""
    import openpyxl

    template_path = tmp_output / "_template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TBKQ"
    ws["A1"] = "THÔNG BÁO PHIẾU LƯƠNG"
    ws["A2"] = "Kỳ lương tháng 11/2025"
    ws["A16"] = "A. TĂNG THU NHẬP TRONG KỲ"
    ws["A17"] = "I. THU NHẬP"
    ws["A18"] = "(1) Lương cơ bản"
    ws["A53"] = "C. THỰC LĨNH (A - B)"
    wb.save(str(template_path))
    wb.close()
    return template_path
