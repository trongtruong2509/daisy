"""
Excel reader for payslip data.

Reads employee data from the Data sheet, email template from bodymail,
and TBKQ template structure using xlrd (.xls) or openpyxl (.xlsx).
"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)


def _col_letter_to_index(col: str) -> int:
    """
    Convert column letter to 0-based index.

    Examples: A->0, B->1, Z->25, AA->26, AZ->51
    """
    return column_index_from_string(col) - 1


def _cell_to_row_col(cell_ref: str) -> Tuple[int, int]:
    """
    Parse cell reference like 'B3' to (row_0based, col_0based).

    Args:
        cell_ref: Cell reference (e.g., 'B3', 'AA10').

    Returns:
        Tuple of (row_index_0based, col_index_0based).
    """
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_str, row_str = match.groups()
    return int(row_str) - 1, _col_letter_to_index(col_str)


class ExcelReader:
    """
    Reads payslip-related data from an Excel file.

    Supports both .xls (via xlrd) and .xlsx (via openpyxl).
    """

    def __init__(self, excel_path: Path):
        self.excel_path = Path(excel_path)
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        self._ext = self.excel_path.suffix.lower()
        self._workbook = None
        self._is_xls = self._ext in (".xls",)

    def open(self):
        """Open the Excel workbook.

        For .xls files with formulas, uses win32com (Excel COM) to
        recalculate and save a temp .xlsx copy, then reads via openpyxl.
        Falls back to xlrd if win32com is not available.
        """
        if self._is_xls:
            # Try win32com first for proper formula evaluation
            if self._try_open_via_com():
                logger.info(f"Opened Excel file (via COM): {self.excel_path}")
                return

            # Fallback to xlrd (formula cells will return None)
            try:
                import xlrd
            except ImportError:
                raise ImportError("xlrd is required for .xls files: pip install xlrd")
            self._workbook = xlrd.open_workbook(str(self.excel_path))
            logger.warning(
                "Opened with xlrd — formula cells will not have computed values"
            )
        else:
            import openpyxl

            self._workbook = openpyxl.load_workbook(
                str(self.excel_path), data_only=True, read_only=True
            )
        logger.info(f"Opened Excel file: {self.excel_path}")

    def _try_open_via_com(self) -> bool:
        """Open .xls via Excel COM, recalculate, and save as temp .xlsx.

        This ensures formula cells have their computed values.

        Returns:
            True if successfully opened via COM, False otherwise.
        """
        try:
            import win32com.client as win32
            import pythoncom
        except ImportError:
            logger.debug("win32com not available, skipping COM reader")
            return False

        try:
            pythoncom.CoInitialize()
            excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            try:
                wb = excel_app.Workbooks.Open(
                    str(self.excel_path.resolve()),
                    ReadOnly=True,
                    UpdateLinks=0,  # Don't update external links
                )

                # Set manual calculation AFTER opening to preserve
                # cached formula values when referenced sheets are missing
                try:
                    excel_app.Calculation = -4135  # xlCalculationManual
                    excel_app.AskToUpdateLinks = False
                except Exception:
                    pass

                # Do NOT call CalculateFull() — preserve cached formula values

                # Save as temp .xlsx so openpyxl can read computed values
                temp_xlsx = self.excel_path.parent / f"_temp_{self.excel_path.stem}.xlsx"
                wb.SaveAs(
                    str(temp_xlsx.resolve()),
                    FileFormat=51,  # xlOpenXMLWorkbook
                )
                wb.Close(SaveChanges=False)
            finally:
                excel_app.Quit()
                pythoncom.CoUninitialize()

            # Now open the temp .xlsx with openpyxl
            import openpyxl
            self._workbook = openpyxl.load_workbook(
                str(temp_xlsx), data_only=True, read_only=True
            )
            self._is_xls = False  # Switch to openpyxl mode for reading
            self._temp_xlsx = temp_xlsx  # Track for cleanup
            return True

        except Exception as e:
            logger.warning(f"COM reader failed, falling back to xlrd: {e}")
            return False

    def close(self):
        """Close the workbook and clean up temp files."""
        if self._workbook and not self._is_xls:
            self._workbook.close()
        self._workbook = None

        # Clean up temp .xlsx if we created one via COM
        temp = getattr(self, "_temp_xlsx", None)
        if temp and Path(temp).exists():
            try:
                Path(temp).unlink()
                logger.debug(f"Cleaned up temp file: {temp}")
            except Exception:
                pass

    def __enter__(self):
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False

    def _get_sheet(self, sheet_name: str):
        """Get a sheet by name."""
        if self._workbook is None:
            raise RuntimeError("Workbook not opened. Call open() first.")
        if self._is_xls:
            return self._workbook.sheet_by_name(sheet_name)
        else:
            if sheet_name not in self._workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            return self._workbook[sheet_name]

    def _get_cell_value(self, sheet, row: int, col: int) -> Any:
        """
        Get cell value by 0-based row and col index.

        Handles error type cells by returning None (including Excel error
        strings like #NAME?, #REF!, etc. and COM error integers).
        """
        if self._is_xls:
            import xlrd

            ctype = sheet.cell_type(row, col)
            if ctype == xlrd.XL_CELL_ERROR:
                return None
            if ctype == xlrd.XL_CELL_EMPTY:
                return None
            value = sheet.cell_value(row, col)
            # Handle dates
            if ctype == xlrd.XL_CELL_DATE:
                try:
                    date_tuple = xlrd.xldate_as_tuple(value, self._workbook.datemode)
                    return value  # Return as float for now
                except Exception:
                    return value
            return self._filter_error_value(value)
        else:
            cell = sheet.cell(row=row + 1, column=col + 1)
            return self._filter_error_value(cell.value)

    @staticmethod
    def _filter_error_value(value: Any) -> Any:
        """Return None for Excel error values (strings or COM integers)."""
        if value is None:
            return None
        # COM error integers (e.g., -2146826259 for #NAME?)
        if isinstance(value, int) and value < -2000000000:
            return None
        # Excel error strings
        if isinstance(value, str) and value.strip().startswith("#"):
            error_strings = {
                "#DIV/0!", "#N/A", "#NAME?", "#NULL!",
                "#NUM!", "#REF!", "#VALUE!", "#GETTING_DATA",
            }
            if value.strip() in error_strings:
                return None
        return value

    def _get_cell_by_ref(self, sheet, cell_ref: str) -> Any:
        """Get cell value by reference like 'A1', 'B3', etc."""
        row, col = _cell_to_row_col(cell_ref)
        return self._get_cell_value(sheet, row, col)

    def read_employees(
        self,
        data_sheet: str,
        header_row: int,
        start_row: int,
        col_mnv: str,
        col_name: str,
        col_email: str,
        col_password: str,
    ) -> List[Dict[str, Any]]:
        """
        Read employee data from the Data sheet.

        Args:
            data_sheet: Name of the Data sheet.
            header_row: 1-based row number of headers.
            start_row: 1-based row number of first employee.
            col_mnv: Column letter for MNV.
            col_name: Column letter for Name.
            col_email: Column letter for Email.
            col_password: Column letter for Password.

        Returns:
            List of employee dicts with keys: mnv, name, email, password,
            and all other columns keyed by column letter.
        """
        sheet = self._get_sheet(data_sheet)

        # Determine dimensions
        if self._is_xls:
            nrows = sheet.nrows
            ncols = sheet.ncols
        else:
            nrows = sheet.max_row or 0
            ncols = sheet.max_column or 0

        # Read headers to build column letter → header name mapping
        header_row_idx = header_row - 1  # 0-based
        headers = {}
        for c in range(ncols):
            val = self._get_cell_value(sheet, header_row_idx, c)
            if val is not None:
                col_letter = self._index_to_col_letter(c)
                headers[col_letter] = str(val).strip()

        logger.info(f"Data sheet: {nrows} rows, {ncols} cols, {len(headers)} headers")

        # Key column indices
        mnv_idx = _col_letter_to_index(col_mnv)
        name_idx = _col_letter_to_index(col_name)
        email_idx = _col_letter_to_index(col_email)
        pw_idx = _col_letter_to_index(col_password)

        employees = []
        start_row_idx = start_row - 1  # 0-based

        for r in range(start_row_idx, nrows):
            mnv_val = self._get_cell_value(sheet, r, mnv_idx)

            # Skip empty rows (no MNV)
            if mnv_val is None or str(mnv_val).strip() == "":
                continue

            # Build employee dict with ALL column data
            emp = {
                "row": r + 1,  # 1-based row number
                "mnv": self._normalize_mnv(mnv_val),
                "name": str(self._get_cell_value(sheet, r, name_idx) or "").strip(),
                "email": str(self._get_cell_value(sheet, r, email_idx) or "").strip(),
                "password": self._normalize_password(
                    self._get_cell_value(sheet, r, pw_idx)
                ),
            }

            # Read all columns and store by column letter
            emp["columns"] = {}
            for c in range(ncols):
                col_letter = self._index_to_col_letter(c)
                val = self._get_cell_value(sheet, r, c)
                emp["columns"][col_letter] = val

            # If Data sheet columns have errors (None values due to XLOOKUP failures),
            # try to lookup values directly from 'bang luong' sheet
            self._fill_from_bang_luong(emp)

            employees.append(emp)

        logger.info(f"Read {len(employees)} employees from {data_sheet}")
        return employees

    def _fill_from_bang_luong(self, employee: Dict[str, Any]):
        """
        Fill missing employee data by looking up in 'bang luong' sheet.
        
        The Data sheet uses XLOOKUP formulas that may not work in all Excel versions.
        This method provides a fallback by directly reading from 'bang luong' sheet.
        
        Maps MNV → values from 'bang luong' columns based on the XLOOKUP formulas:
        - Column B (Name): bang luong column M
        - Column J: bang luong column AE  
        - Column K: bang luong column AF
        - Column N: bang luong column AI
        - etc.
        """
        try:
            # Check if we need to fill (if key columns are None/missing)
            columns = employee.get("columns", {})
            mnv = employee.get("mnv")
            if not mnv:
                return
            
            # Check if we need to lookup (Name is None or error value means XLOOKUP failed)
            name_val = columns.get("B")
            # Error code 29 is #VALUE! error in xlrd
            # If name is None, empty, numeric error code, or not a string, we need lookup
            needs_lookup = (
                name_val is None 
                or name_val == "" 
                or (isinstance(name_val, (int, float)) and name_val in (29, 15,  42, 7))  # Common Excel error codes
                or len(str(name_val).strip()) < 2  # Name too short to be valid
            )
            
            if not needs_lookup:
                # Data sheet values are OK, no need for lookup
                logger.debug(f"Data sheet values OK for MNV {mnv}, skipping bang luong lookup")
                return
            
            logger.debug(f"Data sheet has errors for MNV {mnv}, attempting bang luong lookup")
            
            # Try to read bang luong sheet
            try:
                bang_luong_sheet = self._get_sheet("bang luong")
            except Exception as e:
                # bang luong sheet doesn't exist, can't help
                logger.warning(f"Could not access 'bang luong' sheet: {e}")
                return
            
            # Find the row in bang luong where column L matches MNV
            # Column L is the 12th column (0-indexed: 11)
            if self._is_xls:
                nrows = bang_luong_sheet.nrows
            else:
                nrows = bang_luong_sheet.max_row or 0
            
            matched_row_idx = None
            mnv_str = str(mnv).strip()
            
            for r_idx in range(nrows):
                val = self._get_cell_value(bang_luong_sheet, r_idx, 11)  # Column L (0-indexed: 11)
                if val is not None:
                    val_str = str(val).strip().rstrip('.0')  # Handle 6046072.0 → 6046072
                    if val_str == mnv_str:
                        matched_row_idx = r_idx
                        break
            
            if matched_row_idx is None:
                logger.warning(f"MNV {mnv} not found in 'bang luong' sheet")
                return
            
            # Map bang luong columns to Data sheet columns based on XLOOKUP formulas
            # Data B: =XLOOKUP(A4, L:L, M:M) → bang luong col M (index 12)
            # Data J: =XLOOKUP(A4, L:L, AE:AE) → bang luong col AE (index 30)
            # Data K: =XLOOKUP(A4, L:L, T:T) → bang luong col T (index 19) - NGÀY CÔNG CHUẨN
            # Data N: =XLOOKUP(A4, L:L, V:V) → bang luong col V (index 21) - CÔNG THỰC TẾ
            # Data O: =XLOOKUP(A4, L:L, AJ:AJ) → bang luong col AJ (index 35) - LƯƠNG THEO GIỜ CÔNG
            # Data P: =XLOOKUP(A4, L:L, AO:AO) → bang luong col AO (index 40) - TIỀN LƯƠNG LÀM ĐÊM
            # Data Q: =XLOOKUP(A4, L:L, AM:AM) → bang luong col AM (index 38) - TỔNG LƯƠNG LÀM THÊM GIỜ
            # Data R: =XLOOKUP(A4, L:L, AQ:AQ) → bang luong col AQ (index 42) - THANH TOÁN QUỸ NGHỈ BÙ
            # Data S: =XLOOKUP(A4, L:L, AP:AP) → bang luong col AP (index 41) - THANH TOÁN PHÉP NĂM
            # Data T: =XLOOKUP(A4, L:L, BA:BA) → bang luong col BA (index 52) - HỖ TRỢ CHI PHÍ GỬI XE
            
            lookup_map = {
                "B": 12,   # M (Name)
                "J": 30,   # AE (THU NHẬP CƠ BẢN GROSS - Mức lương)
                "K": 19,   # T (NGÀY CÔNG CHUẨN - Công chuẩn)
                "N": 21,   # V (CÔNG THỰC TẾ - Tổng công)
                "O": 35,   # AJ (LƯƠNG THEO GIỜ CÔNG)
                "P": 40,   # AO (TIỀN LƯƠNG LÀM ĐÊM)
                "Q": 38,   # AM (TỔNG LƯƠNG LÀM THÊM GIỜ _TAX)
                "R": 42,   # AQ (THANH TOÁN QUỸ NGHỈ BÙ)
                "S": 41,   # AP (THANH TOÁN PHÉP NĂM)
                "T": 52,   # BA (HỖ TRỢ CHI PHÍ GỬI XE)
                "U": 38,   # AM (TỔNG LƯƠNG LÀM THÊM GIỜ _TAX - for calculations)
                "V": 54,   # BC (HỖ TRỢ ĐIỆN THOẠI)
                "W": 43,   # AR (PC CHUYÊN MÔN / TAY NGHỀ)
                "X": 44,   # AS (PC KIÊM NHIỆM)
                "Y": 45,   # AT (PC TRÁCH NHIỆM)
                "Z": 46,   # AU (PHỤ CẤP TRÁCH NHIỆM CÔNG VIỆC)
                "AA": 47,  # AV (PHỤ CẤP TRÁCH NHIỆM ĐỀ BẠT)
                "AB": 48,  # AW (TIỀN ĂN GIỮA CA - CHỊU THUẾ)
                "AC": 49,  # AX (TIỀN ĂN GIỮA CA - KHÔNG CHỊU THUẾ)
                "AD": 50,  # AY ( HỖ TRỢ NHÀ Ở)
                "AH": 54,  # BC (HỖ TRỢ ĐIỆN THOẠI - or net payment calculation)
            }
            
            for data_col, bl_col_idx in lookup_map.items():
                if columns.get(data_col) is None:  # Only fill if missing
                    val = self._get_cell_value(bang_luong_sheet, matched_row_idx, bl_col_idx)
                    columns[data_col] = val
            
            # Also update the name field in employee dict
            if columns.get("B"):
                employee["name"] = str(columns["B"]).strip()
            
            logger.debug(f"Filled data for MNV {mnv} from 'bang luong' sheet")
            
        except Exception as e:
            logger.warning(f"Failed to fill from 'bang luong' sheet for MNV {employee.get('mnv')}: {e}")

    def read_email_template(
        self,
        sheet_name: str,
        body_cells: List[str],
        date_cell: str,
    ) -> Dict[str, Any]:
        """
        Read email body template from bodymail sheet.

        Args:
            sheet_name: Name of the bodymail sheet.
            body_cells: List of cell references to read (e.g., ['A1', 'A3']).
            date_cell: Cell containing date placeholder text.

        Returns:
            Dict with cell reference → value mapping.
        """
        sheet = self._get_sheet(sheet_name)

        template = {}
        for cell_ref in body_cells:
            val = self._get_cell_by_ref(sheet, cell_ref)
            template[cell_ref] = str(val) if val is not None else ""

        logger.info(
            f"Read email template from {sheet_name}: "
            f"{len(template)} cells ({', '.join(body_cells)})"
        )
        return template

    def read_email_subject(
        self, sheet_name: str, subject_cell: str
    ) -> str:
        """
        Read email subject from the TBKQ sheet.

        Args:
            sheet_name: Name of the template sheet (TBKQ).
            subject_cell: Cell reference for subject (e.g., 'G1').

        Returns:
            Email subject string.
        """
        sheet = self._get_sheet(sheet_name)
        val = self._get_cell_by_ref(sheet, subject_cell)
        subject = str(val) if val else ""
        logger.info(f"Email subject from {sheet_name}!{subject_cell}: {subject}")
        return subject

    def read_template_structure(
        self, sheet_name: str
    ) -> Dict[str, Any]:
        """
        Read TBKQ template structure for reference (labels, hints, etc.)

        Returns:
            Dict with 'labels' and 'f_column_hints' for debugging.
        """
        sheet = self._get_sheet(sheet_name)

        if self._is_xls:
            nrows = sheet.nrows
        else:
            nrows = sheet.max_row or 0

        labels = {}
        f_hints = {}

        for r in range(nrows):
            # Column A labels
            a_val = self._get_cell_value(sheet, r, 0)
            if a_val is not None:
                labels[f"A{r + 1}"] = str(a_val)

            # Column F hints (mapping references)
            f_val = self._get_cell_value(sheet, r, 5)
            if f_val is not None:
                f_hints[f"F{r + 1}"] = str(f_val)

        return {"labels": labels, "f_hints": f_hints}

    @staticmethod
    def _normalize_mnv(value: Any) -> str:
        """Normalize MNV to string, preserving leading zeros."""
        if value is None:
            return ""
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def _normalize_password(value: Any) -> str:
        """Normalize password: strip leading zeros per spec."""
        if value is None:
            return ""
        if isinstance(value, float) and value == int(value):
            pw = str(int(value))
        else:
            pw = str(value).strip()
        return pw.lstrip("0") or "0"

    @staticmethod
    def _index_to_col_letter(index: int) -> str:
        """Convert 0-based column index to letter(s). 0->A, 25->Z, 26->AA."""
        result = ""
        while True:
            result = chr(65 + (index % 26)) + result
            index = index // 26 - 1
            if index < 0:
                break
        return result
