"""Unit tests for payslip_generator module."""

import sys
from pathlib import Path

import openpyxl
import pytest

TOOL_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(TOOL_DIR))

from payslip_generator import PayslipGenerator, _cell_to_row_col


class TestCellParsing:
    """Test cell reference parsing."""

    def test_simple_cell(self):
        row, col = _cell_to_row_col("A1")
        assert row == 1
        assert col == 1

    def test_multi_letter_cell(self):
        row, col = _cell_to_row_col("AA10")
        assert row == 10
        assert col == 27

    def test_cell_b3(self):
        row, col = _cell_to_row_col("B3")
        assert row == 3
        assert col == 2

    def test_cell_d53(self):
        row, col = _cell_to_row_col("D53")
        assert row == 53
        assert col == 4

    def test_invalid_cell(self):
        with pytest.raises(ValueError):
            _cell_to_row_col("invalid")


class TestPayslipGenerator:
    """Tests for PayslipGenerator."""

    def test_init(self, minimal_template, tmp_output, default_cell_mapping, default_calc_mapping):
        gen = PayslipGenerator(
            template_path=minimal_template,
            output_dir=tmp_output,
            cell_mapping=default_cell_mapping,
            calc_mapping=default_calc_mapping,
            date_str="01/2026",
        )
        assert gen.month == "01"
        assert gen.year == "2026"
        assert gen.mmyyyy == "012026"

    def test_generate_single_payslip(
        self, minimal_template, tmp_output,
        default_cell_mapping, default_calc_mapping,
        sample_employee,
    ):
        gen = PayslipGenerator(
            template_path=minimal_template,
            output_dir=tmp_output,
            cell_mapping=default_cell_mapping,
            calc_mapping=default_calc_mapping,
            date_str="01/2026",
            filename_pattern="TBKQ_{name}_{mmyyyy}.xlsx",
        )
        result = gen.generate_payslip(sample_employee)
        assert result is not None
        assert result.exists()
        assert "Nguyen Van A" in result.name
        assert "012026" in result.name

        # Verify cell values
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active
        assert ws.cell(row=3, column=2).value == 6046072.0  # B3 = MNV
        assert ws.cell(row=4, column=2).value == "Nguyen Van A"  # B4 = Name
        assert ws.cell(row=18, column=4).value == 12000000.0  # D18 = base salary
        assert ws.cell(row=53, column=4).value == 13050000.0  # D53 = net pay
        wb.close()

    def test_calculated_cells(
        self, minimal_template, tmp_output,
        default_cell_mapping, default_calc_mapping,
        sample_employee,
    ):
        gen = PayslipGenerator(
            template_path=minimal_template,
            output_dir=tmp_output,
            cell_mapping=default_cell_mapping,
            calc_mapping=default_calc_mapping,
            date_str="01/2026",
        )
        result = gen.generate_payslip(sample_employee)
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active

        # D17 = D18 = 12000000
        assert ws.cell(row=17, column=4).value == 12000000.0

        # D22 = D24 + D28 = P + T = 2000000 + 1000000 = 3000000
        assert ws.cell(row=22, column=4).value == 3000000.0

        # D38 = U+W+X+Y = 100000+0+0+0 = 100000
        assert ws.cell(row=38, column=4).value == 100000.0

        # D30 = D31+D32+D33+D37+D38 = 500000+300000+200000+500000+100000 = 1600000
        assert ws.cell(row=30, column=4).value == 1600000.0

        # D21 = D22 + D30 = 3000000 + 1600000 = 4600000
        assert ws.cell(row=21, column=4).value == 4600000.0

        # D16 = D17 + D21 = 12000000 + 4600000 = 16600000
        assert ws.cell(row=16, column=4).value == 16600000.0

        # D49 = D50 + D51 = AB + AC = 1200000 + 100000 = 1300000
        assert ws.cell(row=49, column=4).value == 1300000.0

        # D45 = 0
        assert ws.cell(row=45, column=4).value == 0.0

        # D40 = D44 + D45 + D48 = Z + 0 + AA = 200000 + 0 + 150000 = 350000
        assert ws.cell(row=40, column=4).value == 350000.0

        # D39 = D40 + D49 + D52 = 350000 + 1300000 + AD=500000 = 2150000
        assert ws.cell(row=39, column=4).value == 2150000.0

        wb.close()

    def test_empty_name_uses_mnv_in_filename(
        self, minimal_template, tmp_output,
        default_cell_mapping, default_calc_mapping,
        sample_employee,
    ):
        sample_employee["name"] = ""
        gen = PayslipGenerator(
            template_path=minimal_template,
            output_dir=tmp_output,
            cell_mapping=default_cell_mapping,
            calc_mapping=default_calc_mapping,
            date_str="01/2026",
            filename_pattern="TBKQ_{name}_{mmyyyy}.xlsx",
        )
        result = gen.generate_payslip(sample_employee)
        assert result is not None
        assert "6046072" in result.name

    def test_generate_batch(
        self, minimal_template, tmp_output,
        default_cell_mapping, default_calc_mapping,
        sample_employees,
    ):
        gen = PayslipGenerator(
            template_path=minimal_template,
            output_dir=tmp_output,
            cell_mapping=default_cell_mapping,
            calc_mapping=default_calc_mapping,
            date_str="02/2026",
        )
        results = gen.generate_batch(sample_employees)
        assert len(results) == 2
        assert all(r["success"] for r in results)

    def test_date_cells_updated(
        self, minimal_template, tmp_output,
        default_cell_mapping, default_calc_mapping,
        sample_employee,
    ):
        gen = PayslipGenerator(
            template_path=minimal_template,
            output_dir=tmp_output,
            cell_mapping=default_cell_mapping,
            calc_mapping=default_calc_mapping,
            date_str="03/2026",
        )
        result = gen.generate_payslip(sample_employee)
        wb = openpyxl.load_workbook(str(result))
        ws = wb.active
        a2 = ws.cell(row=2, column=1).value
        assert "03/2026" in str(a2)
        wb.close()
