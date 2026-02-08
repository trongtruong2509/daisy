"""
Payslip generator using Excel COM (win32com).

Replicates the VBA macro approach:
  1. Open source workbook with Excel COM
  2. For each employee, set TBKQ!B3 = MNV to trigger INDEX/MATCH formulas
  3. Copy the calculated TBKQ sheet to a new workbook
  4. Paste as values, delete helper columns, clean up
  5. Save as .xlsx

Falls back to openpyxl-based generation when COM is not available.
"""

import logging
import re
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

logger = logging.getLogger(__name__)

# Excel error values returned by COM (VBA error codes)
_EXCEL_ERROR_VALUES = {
    -2146826281,  # #DIV/0!
    -2146826246,  # #N/A
    -2146826259,  # #NAME?
    -2146826288,  # #NULL!
    -2146826252,  # #NUM!
    -2146826265,  # #REF!
    -2146826273,  # #VALUE!
}

# Excel error strings that may appear in cell values
_EXCEL_ERROR_STRINGS = {
    "#DIV/0!", "#N/A", "#NAME?", "#NULL!",
    "#NUM!", "#REF!", "#VALUE!", "#GETTING_DATA",
}


def _is_excel_error(value) -> bool:
    """Check if a value is an Excel error (COM integer or string)."""
    if isinstance(value, int) and value in _EXCEL_ERROR_VALUES:
        return True
    if isinstance(value, str) and value.strip() in _EXCEL_ERROR_STRINGS:
        return True
    return False


def _col_letter_to_index(col: str) -> int:
    """Column letter to 1-based index for openpyxl."""
    return column_index_from_string(col)


def _cell_to_row_col(cell_ref: str):
    """Parse 'B3' to (row_1based, col_1based) for openpyxl."""
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref.upper())
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_str, row_str = match.groups()
    return int(row_str), _col_letter_to_index(col_str)


class PayslipGenerator:
    """
    Generates payslip Excel files from a template.

    Flow:
    1. Use a prepared .xlsx template (TBKQ sheet only)
    2. For each employee, copy template and fill data directly
    3. Save as .xlsx (later converted to PDF)
    """

    def __init__(
        self,
        template_path: Path,
        output_dir: Path,
        cell_mapping: Dict[str, str],
        calc_mapping: Dict[str, str],
        date_str: str,
        filename_pattern: str = "TBKQ_{name}_{mmyyyy}.xlsx",
    ):
        """
        Args:
            template_path: Path to TBKQ template .xlsx file.
            output_dir: Directory for generated payslip files.
            cell_mapping: TBKQ cell -> Data column mapping (e.g., {"B3": "A"}).
            calc_mapping: Calculated cell formulas (e.g., {"D16": "=D17+D21"}).
            date_str: Payroll date in MM/YYYY format.
            filename_pattern: Pattern for output filenames.
        """
        self.template_path = Path(template_path)
        self.output_dir = Path(output_dir)
        self.cell_mapping = cell_mapping
        self.calc_mapping = calc_mapping
        self.date_str = date_str
        self.filename_pattern = filename_pattern

        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Parse date
        parts = date_str.split("/")
        self.month = parts[0] if len(parts) >= 1 else ""
        self.year = parts[1] if len(parts) >= 2 else ""
        self.mmyyyy = self.month + self.year

    def prepare_template(
        self,
        source_xls: Path,
        template_sheet: str = "TBKQ",
    ) -> Path:
        """
        Prepare a clean .xlsx template from the source .xls file.

        Uses win32com (Excel COM) to:
        1. Open the .xls file, recalculate all formulas
        2. Copy the TBKQ sheet to a new workbook
        3. Paste as values (while source is still open for formula resolution)
        4. Delete helper columns F/G, clean up buttons
        5. Clear cells that will be filled dynamically per employee
        6. Save as .xlsx template

        This is a one-time operation per run.

        Args:
            source_xls: Path to source Excel file (.xls or .xlsx).
            template_sheet: Sheet name to extract.

        Returns:
            Path to the prepared template .xlsx file.
        """
        template_out = self.output_dir / "_template.xlsx"

        if template_out.exists():
            logger.info(f"Template already prepared: {template_out}")
            self.template_path = template_out
            return template_out

        logger.info(f"Preparing template from {source_xls}...")

        try:
            import win32com.client as win32
            import pythoncom

            pythoncom.CoInitialize()

            excel = win32.gencache.EnsureDispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False

            try:
                wb = excel.Workbooks.Open(str(source_xls.resolve()))

                # Recalculate all formulas while all sheets are present
                excel.CalculateFull()

                ws = wb.Sheets(template_sheet)

                # Copy sheet to new workbook
                ws.Copy()
                new_wb = excel.ActiveWorkbook
                new_ws = new_wb.ActiveSheet

                # Paste as values FIRST — while source workbook is still open
                # so cross-sheet formula results are captured correctly
                new_ws.Cells.Copy()
                new_ws.Cells.PasteSpecial(Paste=-4163)  # xlPasteValues
                excel.CutCopyMode = False

                # Now delete columns F and G (mapping hints, not needed)
                # Column G first (higher index), then F
                new_ws.Columns("G").Delete()
                new_ws.Columns("F").Delete()

                # Clear cells that will be filled dynamically per employee
                cells_to_clear = (
                    list(self.cell_mapping.keys())
                    + list(self.calc_mapping.keys())
                )
                for cell_ref in cells_to_clear:
                    try:
                        new_ws.Range(cell_ref).ClearContents()
                    except Exception:
                        pass

                # Clear cells used by macro buttons (originally K2:L2,
                # shifted to I2:J2 after column deletion)
                try:
                    new_ws.Range("I2", "J2").ClearContents()
                except Exception:
                    pass

                # Delete any buttons
                try:
                    for btn in new_ws.Buttons:
                        btn.Delete()
                except Exception:
                    pass

                # Set print area
                new_ws.PageSetup.PrintArea = "$A$1:$E$61"

                # Save as xlsx (no password)
                new_wb.SaveAs(
                    str(template_out.resolve()),
                    FileFormat=51,  # xlOpenXMLWorkbook
                )
                new_wb.Close(SaveChanges=False)

            finally:
                wb.Close(SaveChanges=False)
                excel.Quit()
                pythoncom.CoUninitialize()

            logger.info(f"Template prepared: {template_out}")
            self.template_path = template_out
            return template_out

        except ImportError:
            logger.warning(
                "win32com not available. Creating template from scratch with openpyxl."
            )
            return self._prepare_template_openpyxl(source_xls, template_sheet)

    def generate_batch_via_com(
        self,
        employees: List[Dict[str, Any]],
        source_xls: Path,
        template_sheet: str = "TBKQ",
    ) -> List[Dict[str, Any]]:
        """
        Generate payslips by replicating the VBA macro approach.

        For each employee:
        1. Set TBKQ!B3 = MNV (triggers INDEX/MATCH recalculation)
        2. Copy the TBKQ sheet to a new workbook
        3. Paste as values, delete columns F/G, clean up
        4. Save as .xlsx with employee-specific name

        This produces payslips with all values correctly resolved,
        even when Data sheet formulas reference external sheets.

        Args:
            employees: List of employee data dicts.
            source_xls: Path to the source Excel file.
            template_sheet: Name of the TBKQ sheet.

        Returns:
            List of result dicts with 'employee', 'xlsx_path', 'success'.
        """
        import time
        import win32com.client as win32
        import pythoncom

        results = []
        total = len(employees)

        # Allow any previous COM session to fully release
        time.sleep(2)

        pythoncom.CoInitialize()

        # Use Dispatch (not gencache.EnsureDispatch) to avoid cache issues
        # after a previous COM session in the same process.
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        try:
            wb = excel.Workbooks.Open(
                str(source_xls.resolve()),
                UpdateLinks=0,  # Don't update external links
            )

            # Set manual calculation AFTER opening to prevent
            # unnecessary recalculation during payslip generation
            try:
                excel.Calculation = -4135  # xlCalculationManual
            except Exception:
                pass  # Ignore if not available
            excel.AskToUpdateLinks = False

            for i, emp in enumerate(employees, 1):
                mnv = emp.get("mnv", "")
                name = emp.get("name", "")
                password = emp.get("password", "")

                logger.info(
                    f"[{i}/{total}] Generating payslip for {name} (MNV: {mnv})"
                )

                try:
                    xlsx_path = self._generate_one_via_com(
                        excel, wb, template_sheet, emp
                    )
                    results.append({
                        "employee": emp,
                        "xlsx_path": xlsx_path,
                        "success": xlsx_path is not None,
                    })
                except Exception as e:
                    logger.error(
                        f"Failed to generate payslip for {name}: {e}"
                    )
                    results.append({
                        "employee": emp,
                        "xlsx_path": None,
                        "success": False,
                    })

            wb.Close(SaveChanges=False)
        finally:
            excel.Quit()
            pythoncom.CoUninitialize()

        success = sum(1 for r in results if r["success"])
        failed = total - success
        logger.info(
            f"Payslip generation complete: {success} success, {failed} failed"
        )
        return results

    def _generate_one_via_com(
        self,
        excel,
        source_wb,
        template_sheet: str,
        employee: Dict[str, Any],
    ) -> Optional[Path]:
        """
        Generate a single payslip by setting B3 and copying the TBKQ sheet.

        Replicates VBA logic:
          Range("B3") = Data.Range("A" & i)  ' set MNV
          Worksheets("TBKQ").Copy             ' copy to new workbook
          ... paste values, delete cols, save
        """
        mnv = employee.get("mnv", "")
        name = employee.get("name", "")

        # Generate output filename
        safe_name = re.sub(r'[\\/*?:"<>|]', "_", name) if name else mnv
        filename = self.filename_pattern.replace(
            "{name}", safe_name
        ).replace("{mmyyyy}", self.mmyyyy)
        xlsx_filename = filename.replace(".pdf", ".xlsx")
        output_path = self.output_dir / xlsx_filename

        # Step 1: Set TBKQ!B3 = MNV to trigger INDEX/MATCH formulas
        ws = source_wb.Sheets(template_sheet)
        ws.Range("B3").Value = mnv

        # Step 2: Recalculate ONLY the TBKQ sheet (not the entire workbook).
        # This preserves cached values in the Data sheet while resolving
        # TBKQ formulas (INDEX/MATCH) against those cached values.
        ws.Calculate()

        # Step 3: Copy the TBKQ sheet to a new workbook
        ws.Copy()
        new_wb = excel.ActiveWorkbook
        new_ws = new_wb.ActiveSheet

        # Step 4: Delete buttons
        try:
            new_ws.Buttons.Delete()
        except Exception:
            pass

        # Step 5: Clear macro button cells (K2:L2 in original)
        try:
            new_ws.Range("K2", "L2").ClearContents()
        except Exception:
            pass

        # Step 6: Delete columns F and G
        new_ws.Columns("G").Delete()
        new_ws.Columns("F").Delete()

        # Step 7: Paste as values to freeze all formula results
        new_ws.Cells.Copy()
        new_ws.Cells.PasteSpecial(Paste=-4163)  # xlPasteValues
        excel.CutCopyMode = False

        # Step 8: Delete named ranges from the new workbook
        try:
            for named in list(new_wb.Names):
                named.Delete()
        except Exception:
            pass

        # Step 9: Set print area
        new_ws.PageSetup.PrintArea = "$A$1:$E$61"

        # Step 10: Update date in A2
        a2_val = new_ws.Range("A2").Value
        if a2_val and isinstance(a2_val, str) and self.month and self.year:
            updated = re.sub(
                r"tháng\s+\d{1,2}/\d{4}",
                f"tháng {self.month}/{self.year}",
                a2_val,
            )
            new_ws.Range("A2").Value = updated

        # Step 11: Save as .xlsx
        new_wb.SaveAs(
            str(output_path.resolve()),
            FileFormat=51,  # xlOpenXMLWorkbook
        )
        new_wb.Close(SaveChanges=False)

        logger.debug(f"Generated payslip for {name} (MNV: {mnv}): {output_path}")
        return output_path

    def _prepare_template_openpyxl(
        self, source_xls: Path, template_sheet: str
    ) -> Path:
        """
        Fallback: prepare template using openpyxl only (for .xlsx sources).

        Limited: won't work with .xls files, won't preserve all formatting.
        """
        template_out = self.output_dir / "_template.xlsx"

        if source_xls.suffix.lower() == ".xls":
            raise RuntimeError(
                f"Cannot process .xls file without win32com. "
                f"Please convert {source_xls} to .xlsx first, or install pywin32."
            )

        wb = load_workbook(str(source_xls))
        # Keep only the template sheet
        for name in wb.sheetnames:
            if name != template_sheet:
                del wb[name]

        ws = wb.active
        # Clear formula cells (replace with None)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = None

        wb.save(str(template_out))
        wb.close()

        logger.info(f"Template prepared (openpyxl fallback): {template_out}")
        self.template_path = template_out
        return template_out

    def generate_payslip(
        self,
        employee: Dict[str, Any],
    ) -> Optional[Path]:
        """
        Generate a single payslip Excel file.

        Args:
            employee: Employee data dict with 'columns' sub-dict.

        Returns:
            Path to generated .xlsx file, or None on failure.
        """
        mnv = employee.get("mnv", "")
        name = employee.get("name", "")
        columns = employee.get("columns", {})

        # Generate output filename
        safe_name = re.sub(r'[\\/*?:"<>|]', "_", name) if name else mnv
        filename = self.filename_pattern.replace("{name}", safe_name).replace(
            "{mmyyyy}", self.mmyyyy
        )
        # Replace .pdf with .xlsx for the intermediate file
        xlsx_filename = filename.replace(".pdf", ".xlsx")
        output_path = self.output_dir / xlsx_filename

        try:
            # Copy template
            shutil.copy2(str(self.template_path), str(output_path))

            # Open and fill
            wb = load_workbook(str(output_path))
            ws = wb.active

            # Debug: Check columns data
            non_null_cols = {k: v for k, v in columns.items() if v is not None and v != 0 and v != 0.0 and v != ''}
            logger.info(f"Filling payslip - non-null columns: {list(non_null_cols.keys())}")

            # Step 1: Fill direct mappings from Data columns
            filled_cells = {}
            for tbkq_cell, data_col in self.cell_mapping.items():
                value = columns.get(data_col)
                if value is not None:
                    row, col = _cell_to_row_col(tbkq_cell)
                    ws.cell(row=row, column=col, value=value)
                    filled_cells[tbkq_cell] = value

            # Step 2: Calculate computed cells
            self._fill_calculated_cells(ws, filled_cells, columns)

            # Step 3: Update date-dependent cells
            self._update_date_cells(ws)

            # Step 4: Clean up any Excel error strings that weren't overwritten
            # These come from the template when Data sheet formulas had errors
            self._clear_error_strings(ws)

            wb.save(str(output_path))
            wb.close()

            logger.debug(
                f"Generated payslip for {name} (MNV: {mnv}): {output_path}"
            )
            return output_path

        except Exception as e:
            logger.error(f"Failed to generate payslip for {name}: {e}")
            if output_path.exists():
                output_path.unlink()
            return None

    def _fill_calculated_cells(
        self,
        ws,
        filled_cells: Dict[str, Any],
        columns: Dict[str, Any],
    ):
        """
        Fill calculated cells based on calc_mapping.

        Formulas:
        - "COL1+COL2+..." : Sum of Data columns (e.g., "U+W+X+Y")
        - "=CELL1+CELL2" : Sum of other TBKQ cells (e.g., "=D17+D21")
        - "0" or number : Direct value
        """
        # First pass: calculate cells that reference Data columns directly
        for tbkq_cell, formula in self.calc_mapping.items():
            if formula.startswith("="):
                # References other TBKQ cells — handle in second pass
                continue
            if formula.replace(".", "").isdigit():
                # Direct numeric value
                row, col = _cell_to_row_col(tbkq_cell)
                value = float(formula)
                ws.cell(row=row, column=col, value=value)
                filled_cells[tbkq_cell] = value
                continue

            # Sum of Data columns (e.g., "U+W+X+Y")
            col_refs = [c.strip() for c in formula.split("+")]
            total = 0.0
            for col_ref in col_refs:
                val = columns.get(col_ref)
                if val is not None:
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass

            row, col = _cell_to_row_col(tbkq_cell)
            ws.cell(row=row, column=col, value=total)
            filled_cells[tbkq_cell] = total

        # Second pass: calculate cells that reference other TBKQ cells
        # May need multiple passes for dependent chains
        max_passes = 5
        for _ in range(max_passes):
            unresolved = False
            for tbkq_cell, formula in self.calc_mapping.items():
                if not formula.startswith("="):
                    continue
                if tbkq_cell in filled_cells:
                    continue

                # Parse "=D17+D21" → ["D17", "D21"]
                cell_refs = [c.strip() for c in formula[1:].split("+")]

                # Check if all referenced cells are resolved
                all_resolved = all(ref in filled_cells for ref in cell_refs)
                if not all_resolved:
                    unresolved = True
                    continue

                total = sum(
                    float(filled_cells.get(ref, 0) or 0) for ref in cell_refs
                )
                row, col = _cell_to_row_col(tbkq_cell)
                ws.cell(row=row, column=col, value=total)
                filled_cells[tbkq_cell] = total

            if not unresolved:
                break

    def _update_date_cells(self, ws):
        """Update date-dependent cells in the payslip."""
        if not self.month or not self.year:
            return

        # G1 → title with month (but after F/G deletion, this is now E1)
        # Actually, the template preparation deletes columns F and G,
        # so what was G1 is now E1. However, the title cell reference in
        # config still says "G1". We need to handle this.
        #
        # IMPORTANT: After template prep deletes cols F & G from the template,
        # the remaining columns shift. But we write to the TEMPLATE copy which
        # already has the deletion applied. So we should NOT adjust here —
        # the template already has the right column layout.
        #
        # The title and info cells (A2, A10) are in column A, so they don't shift.
        # We'll update them with the correct month/year.

        # A2 - Info subtitle (update month/year reference)
        a2_val = ws.cell(row=2, column=1).value
        if a2_val and isinstance(a2_val, str):
            # Replace month/year references like "tháng 11/2025"
            updated = re.sub(
                r"tháng\s+\d{1,2}/\d{4}",
                f"tháng {self.month}/{self.year}",
                a2_val,
            )
            ws.cell(row=2, column=1, value=updated)

    def _clear_error_strings(self, ws):
        """
        Clear Excel error strings from worksheet cells.
        
        When the template is created from a source file with formula errors,
        paste-as-values converts those errors to error strings like '#NAME?',
        '#N/A', etc. This method clears those strings.
        """
        error_strings = _EXCEL_ERROR_STRINGS
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in error_strings:
                    cell.value = None

    def generate_batch(
        self, employees: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Generate payslips for all employees.

        Args:
            employees: List of employee data dicts.

        Returns:
            List of result dicts with 'employee', 'xlsx_path', 'success'.
        """
        results = []
        total = len(employees)

        for i, emp in enumerate(employees, 1):
            name = emp.get("name", "N/A")
            mnv = emp.get("mnv", "N/A")

            logger.info(f"[{i}/{total}] Generating payslip for {name} (MNV: {mnv})")

            xlsx_path = self.generate_payslip(emp)

            results.append(
                {
                    "employee": emp,
                    "xlsx_path": xlsx_path,
                    "success": xlsx_path is not None,
                }
            )

        success = sum(1 for r in results if r["success"])
        failed = total - success
        logger.info(
            f"Payslip generation complete: {success} success, {failed} failed"
        )
        return results
